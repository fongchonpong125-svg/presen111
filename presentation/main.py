#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
全网舆情与跨平台热点追踪分析系统
================================================

【使用说明（请仔细阅读并按需修改）】

你需要根据自己的实际情况，修改以下几个位置的配置：

1. API 相关配置（大约在本文件靠前的位置）：
   - 第 1 处：`API_CONFIG` 字典中的三个 URL，如果未来接口有变更，可在这里修改。
   - 第 2 处：`API_CONFIG["common_headers"]["Authorization"]` 中的 Bearer Token（当前已为你填入）。

2. 字体相关配置（用于解决中文乱码）：
   - 第 3 处：`CHINESE_FONTS` 列表与 `get_available_font` 函数，如果你的系统没有这些字体，
     请根据自己电脑实际安装的字体路径进行修改。

其他逻辑你可以按需自行扩展或精简。
本脚本为一个单文件、可直接运行的 Demo 项目。
"""

import os
import sys
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import io
import base64
import webbrowser
import threading
import time
import json
from pathlib import Path

import requests
import pandas as pd
import jieba
from collections import Counter, defaultdict

import matplotlib
import matplotlib.pyplot as plt
from matplotlib import font_manager
from wordcloud import WordCloud

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

from flask import Flask, render_template_string, send_from_directory, abort, request


############################
# 全局基础配置区域（可按需修改）
############################

# 第 1 处 & 第 2 处：API 地址与 API Key 配置
API_CONFIG: Dict[str, Any] = {
    "base_url": "https://v2.xxapi.cn/api",
    # 如果接口路径有变动，只需要修改这里的三个相对路径
    "endpoints": {
        "weibo": "/weibohot",
        "douyin": "/douyinhot",
        "baidu": "/baiduhot",
        "bilibili": "/bilibilihot",
        "kr36": "/hot36kr",
    },
    # 公共请求头（这里演示通过 headers 传递 API Key）
    "common_headers": {
        # ⚠️ 第 2 处：如果你有新的 API Key，请修改下行 Authorization 字段
        "Authorization": "Bearer 54ca8da7dc8dde72",
        "User-Agent": "HotRankAnalyzer/1.0 (+https://example.com)",
        "Accept": "application/json",
    },
    # 公共请求参数（如果未来接口需要在 params 里传递 key，可在这里演示）
    "common_params": {
        # "apikey": "YOUR_API_KEY_IF_NEEDED_IN_PARAMS"
    },
    # 请求相关的超时设置（秒）
    "timeout": 10,
    # 请求重试设置（建议保留，提升稳定性）
    "max_retries": 3,
    "backoff_base_seconds": 0.8,
}

# 平台配置：用于展示、趋势图、网页筛选等（新增平台只需在这里扩展）
PLATFORMS: List[Dict[str, str]] = [
    {"key": "weibo", "name": "微博"},
    {"key": "douyin", "name": "抖音"},
    {"key": "baidu", "name": "百度"},
    {"key": "bilibili", "name": "哔哩哔哩"},
    {"key": "kr36", "name": "36氪"},
]


# 第 3 处：中文字体配置（用于 Matplotlib / WordCloud 防止中文乱码）
CHINESE_FONTS = [
    r"C:\Windows\Fonts\simhei.ttf",        # 常见中文黑体（SimHei）
    r"C:\Windows\Fonts\msyh.ttc",          # 微软雅黑
    r"/System/Library/Fonts/STHeiti Medium.ttc",  # macOS 示意
    r"/usr/share/fonts/truetype/arphic/ukai.ttc",  # Linux 示意
]

# 同义词/别名归一化（可按需扩展）
# 说明：将多个写法映射到同一个“标准词”，可以提升词频、共现统计质量
# 你也可以在项目目录创建 `synonyms.json` 来覆盖/扩展这份映射：
# 格式示例：{"AI": ["人工智能","人工智慧","A.I."], "NBA": ["美职篮"]}
DEFAULT_SYNONYM_MAP: Dict[str, List[str]] = {
    "人工智能": ["AI", "人工智慧", "A.I.", "ai"],
    "新冠": ["新冠病毒", "新冠肺炎", "COVID", "COVID-19", "covid"],
    "美国": ["美國", "USA", "U.S.", "United States"],
}


def load_synonym_map() -> Dict[str, str]:
    """
    加载同义词映射，返回 {别名: 标准词} 的扁平映射。
    优先从项目目录的 `synonyms.json` 读取（若存在），否则使用内置默认映射。
    """
    mapping_source = DEFAULT_SYNONYM_MAP
    path = Path("synonyms.json")
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                mapping_source = data  # type: ignore[assignment]
        except Exception:
            # 读取失败就继续用默认映射
            pass

    flat: Dict[str, str] = {}
    # 支持两种写法：
    # 1) {"标准词": ["别名1","别名2"]}
    # 2) {"别名": "标准词"}
    for k, v in mapping_source.items():
        if isinstance(v, list):
            std = str(k).strip()
            if not std:
                continue
            flat[std] = std
            for alias in v:
                a = str(alias).strip()
                if a:
                    flat[a] = std
        else:
            alias = str(k).strip()
            std = str(v).strip()
            if alias and std:
                flat[alias] = std
                flat[std] = std

    return flat


SYNONYM_FLAT_MAP: Dict[str, str] = load_synonym_map()


def normalize_keyword(word: str) -> str:
    """
    将词语做归一化（同义词合并 + 去空白）。
    """
    w = str(word).strip()
    if not w:
        return ""
    return SYNONYM_FLAT_MAP.get(w, w)


def get_available_font() -> Optional[str]:
    """
    尝试在预设路径中找到第一个存在的中文字体文件。

    返回:
        可用字体路径，若未找到则返回 None。
    """
    for font_path in CHINESE_FONTS:
        if os.path.exists(font_path):
            return font_path
    return None


def setup_logging() -> None:
    """
    初始化日志配置，方便在终端中观察运行情况。
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )


def setup_matplotlib(font_path: Optional[str]) -> None:
    """
    配置 Matplotlib 使用中文字体，避免出现乱码或方块字。

    参数:
        font_path: 字体文件路径，如果为 None 则尝试使用 SimHei 等别名。
    """
    if font_path:
        # 使用字體檔案註冊對應字體名稱，再全域套用
        try:
            font_prop = font_manager.FontProperties(fname=font_path)
            font_name = font_prop.get_name()
            matplotlib.rcParams["font.family"] = font_name
        except Exception:
            # 如果註冊失敗，退回使用常見中文字體名稱
            matplotlib.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "STHeiti"]
    else:
        # 未提供具體字體檔案時，直接指定常見中文字體名稱
        matplotlib.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "STHeiti"]
    # 解决负号 '-' 显示为方块的问题
    matplotlib.rcParams["axes.unicode_minus"] = False


############################
# 模块一：统一数据获取接口
############################

class HotRankFetcher:
    """
    统一封装微博 / 抖音 / 百度热榜的获取逻辑。

    本类的设计目标：
    - 提供一个通用的 GET 请求入口，处理超时和状态码错误。
    - 对外暴露三个方法：`fetch_weibo`, `fetch_douyin`, `fetch_baidu`。
    """

    def __init__(self, config: Dict[str, Any]) -> None:
        self.base_url = config.get("base_url", "").rstrip("/")
        self.endpoints = config.get("endpoints", {})
        self.common_headers = config.get("common_headers", {}).copy()
        self.common_params = config.get("common_params", {}).copy()
        self.timeout = config.get("timeout", 10)
        self.max_retries = int(config.get("max_retries", 1))
        self.backoff_base_seconds = float(config.get("backoff_base_seconds", 0.5))
        # 记录每个平台最近一次请求的状态，供 Web 界面展示
        self.last_request_status: Dict[str, Dict[str, Any]] = {}

    def _build_url(self, platform_key: str) -> str:
        """
        根据平台标识拼接完整 URL。
        """
        endpoint = self.endpoints.get(platform_key, "")
        return f"{self.base_url}{endpoint}"

    def _request(
        self,
        platform_key: str,
        extra_params: Optional[Dict[str, Any]] = None,
        extra_headers: Optional[Dict[str, str]] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        通用 GET 请求方法，包含错误处理。

        参数:
            platform_key: 平台标识（如 "weibo"、"douyin"、"baidu"）
            extra_params: 额外的查询参数，将合并到 common_params 中
            extra_headers: 额外的请求头，将合并到 common_headers 中

        返回:
            成功时返回 JSON 对象（字典），失败时返回 None。
        """
        url = self._build_url(platform_key)
        params = self.common_params.copy()
        headers = self.common_headers.copy()

        if extra_params:
            params.update(extra_params)
        if extra_headers:
            headers.update(extra_headers)

        logging.info("正在请求 %s 热榜: %s", platform_key, url)

        last_error: Optional[str] = None
        status_code: Optional[int] = None
        started_at = time.time()

        for attempt in range(1, max(1, self.max_retries) + 1):
            try:
                response = requests.get(url, headers=headers, params=params, timeout=self.timeout)
                status_code = response.status_code
            except requests.Timeout:
                last_error = "请求超时"
                status_code = None
            except requests.RequestException as e:
                last_error = f"请求异常: {e}"
                status_code = None
            else:
                if response.status_code != 200:
                    last_error = f"状态码异常: {response.status_code}"
                else:
                    try:
                        data = response.json()
                    except ValueError:
                        last_error = "响应无法解析为 JSON"
                    else:
                        code = data.get("code", 200)
                        if code != 200:
                            last_error = f"接口返回错误 code={code}, msg={data.get('msg')}"
                        else:
                            elapsed_ms = int((time.time() - started_at) * 1000)
                            self.last_request_status[platform_key] = {
                                "ok": True,
                                "attempts": attempt,
                                "status_code": response.status_code,
                                "elapsed_ms": elapsed_ms,
                                "error": None,
                            }
                            return data

            # 失败：记录日志并退避重试
            logging.warning(
                "%s 请求失败（第 %d/%d 次）：%s",
                platform_key,
                attempt,
                max(1, self.max_retries),
                last_error,
            )
            if attempt < max(1, self.max_retries):
                sleep_s = self.backoff_base_seconds * (2 ** (attempt - 1))
                time.sleep(sleep_s)

        elapsed_ms = int((time.time() - started_at) * 1000)
        self.last_request_status[platform_key] = {
            "ok": False,
            "attempts": max(1, self.max_retries),
            "status_code": status_code,
            "elapsed_ms": elapsed_ms,
            "error": last_error,
        }
        logging.error("%s 最终失败：%s", platform_key, last_error)
        return None

    def fetch_weibo(self) -> Optional[Dict[str, Any]]:
        """
        获取微博热榜原始 JSON 数据。
        """
        return self._request("weibo")

    def fetch_douyin(self) -> Optional[Dict[str, Any]]:
        """
        获取抖音热榜原始 JSON 数据。
        """
        return self._request("douyin")

    def fetch_baidu(self) -> Optional[Dict[str, Any]]:
        """
        获取百度热榜原始 JSON 数据。
        """
        return self._request("baidu")

    def fetch_bilibili(self) -> Optional[Dict[str, Any]]:
        """
        获取哔哩哔哩热榜原始 JSON 数据。
        """
        return self._request("bilibili")

    def fetch_36kr(self) -> Optional[Dict[str, Any]]:
        """
        获取 36氪 热榜原始 JSON 数据。
        """
        return self._request("kr36")


############################
# 模块二：数据清洗与结构统一
############################

def parse_heat_value(raw: Any) -> float:
    """
    将不同格式的“热度指数”转换为纯数字（float）。

    兼容的常见格式示例：
        - "59万"  -> 590000
        - "3.4亿" -> 340000000
        - "12345" -> 12345
        - 98765   -> 98765
    """
    if raw is None:
        return 0.0

    # 先转换为字符串，去掉空格和逗号
    s = str(raw).strip().replace(",", "")
    if not s:
        return 0.0

    multiplier = 1.0
    if s.endswith("万"):
        multiplier = 1e4
        s = s[:-1]
    elif s.endswith("亿"):
        multiplier = 1e8
        s = s[:-1]
    elif s.lower().endswith("w"):
        multiplier = 1e4
        s = s[:-1]
    elif s.lower().endswith("k"):
        multiplier = 1e3
        s = s[:-1]

    try:
        value = float(s)
    except ValueError:
        # 无法解析则直接返回 0
        return 0.0

    return value * multiplier


def normalize_records_from_json(
    platform_name: str,
    raw_json: Optional[Dict[str, Any]],
) -> pd.DataFrame:
    """
    将单个平台的 JSON 数据解析为统一结构的 DataFrame。

    统一后的字段：
        - Platform: 平台名称（微博 / 抖音 / 百度）
        - Rank:    排名（int）
        - Title:   标题（str）
        - HeatScore: 热度指数（float）

    由于不同平台的 JSON 字段可能不同，这里做了“尽量兼容”的写法：
        - 排名字段优先从 index / rank / sort / hot_index 中尝试；
        - 标题字段优先从 title / word / name 中尝试；
        - 热度字段优先从 hot / hot_value / heat / heat_value 中尝试。
    """
    if not raw_json:
        return pd.DataFrame(columns=["Platform", "Rank", "Title", "HeatScore"])

    data_list = raw_json.get("data") or raw_json.get("list") or raw_json.get("result")
    if not isinstance(data_list, list):
        logging.warning("%s 平台返回的数据结构中未找到列表字段，已跳过。", platform_name)
        return pd.DataFrame(columns=["Platform", "Rank", "Title", "HeatScore"])

    normalized_rows: List[Dict[str, Any]] = []

    def _get_nested(d: Dict[str, Any], keys: List[str]) -> Any:
        cur: Any = d
        for k in keys:
            if not isinstance(cur, dict):
                return None
            cur = cur.get(k)
        return cur

    for idx, item in enumerate(data_list, start=1):
        if not isinstance(item, dict):
            continue

        # 提取排名
        rank = (
            item.get("index")
            or item.get("rank")
            or item.get("sort")
            or item.get("hot_index")
            or idx
        )
        try:
            rank_int = int(rank)
        except Exception:
            rank_int = idx

        # 提取标题（兼容部分平台的嵌套结构，如 36氪 templateMaterial.widgetTitle）
        title = (
            item.get("title")
            or item.get("word")
            or item.get("name")
            or item.get("keyword")
            or _get_nested(item, ["templateMaterial", "widgetTitle"])
            or _get_nested(item, ["templateMaterial", "title"])
            or ""
        )
        title = str(title).strip()
        if not title:
            # 如果连标题都没有，就跳过该条
            continue

        # 提取热度
        raw_heat = (
            item.get("hot")
            or item.get("hot_value")
            or item.get("heat")
            or item.get("heat_value")
            or item.get("hotScore")
            # 36氪：常见为阅读量/点赞等统计字段（尽量取一个能代表热度的数字）
            or _get_nested(item, ["templateMaterial", "statRead"])
            or _get_nested(item, ["templateMaterial", "statPraise"])
            or _get_nested(item, ["templateMaterial", "statCollect"])
        )
        heat_score = parse_heat_value(raw_heat)

        normalized_rows.append(
            {
                "Platform": platform_name,
                "Rank": rank_int,
                "Title": title,
                "HeatScore": heat_score,
            }
        )

    df = pd.DataFrame(normalized_rows, columns=["Platform", "Rank", "Title", "HeatScore"])
    return df


def merge_all_platforms(
    weibo_df: pd.DataFrame,
    douyin_df: pd.DataFrame,
    baidu_df: pd.DataFrame,
    bilibili_df: pd.DataFrame,
    kr36_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    合并三个平台的数据为一个全局 DataFrame。
    """
    all_df = pd.concat([weibo_df, douyin_df, baidu_df, bilibili_df, kr36_df], ignore_index=True)
    # 为了安全起见，删除完全重复的行
    all_df = all_df.drop_duplicates(subset=["Platform", "Rank", "Title"])
    # 按平台 + 排名排序
    all_df = all_df.sort_values(by=["Platform", "Rank"], ascending=[True, True])
    return all_df


############################
# 模块三：跨平台文本分析与共现统计
############################

def load_stopwords() -> set:
    """
    简单内置的中文停用词集合。

    实际项目中可以从文件加载更丰富的停用词表。
    """
    stopwords = {
        "的",
        "了",
        "和",
        "是",
        "在",
        "就",
        "都",
        "而",
        "及",
        "与",
        "着",
        "或",
        "一个",
        "没有",
        "我们",
        "你们",
        "他们",
        "还有",
        "可以",
        "正在",
        "如何",
    }
    # 常见中文/英文标点也一并过滤
    punctuations = set("，。？！：；、“”‘’（）()【】[]<>《》—-…,.!?;:\"'()[]{}<>")
    stopwords |= punctuations
    return stopwords


def tokenize_titles_by_platform(all_df: pd.DataFrame) -> Tuple[Dict[str, List[str]], Counter]:
    """
    使用 jieba 对所有标题进行分词，并按平台分别统计。

    返回:
        - platform_tokens: {平台名: [词1, 词2, ...]}，用于跨平台重合度分析。
        - global_counter: 所有平台合并后的词频 Counter。
    """
    stopwords = load_stopwords()
    platform_tokens: Dict[str, List[str]] = defaultdict(list)
    global_counter: Counter = Counter()

    for _, row in all_df.iterrows():
        platform = row["Platform"]
        title = str(row["Title"])

        # jieba 精确模式分词
        words = jieba.lcut(title)
        for w in words:
            w = w.strip()
            # 过滤长度为 0 的字符串、停用词、纯数字、单个汉字（避免关键词被拆得过碎）
            if not w or w in stopwords:
                continue
            if w.isdigit():
                continue
            if len(w) <= 1:
                continue
            w_norm = normalize_keyword(w)
            if not w_norm or w_norm in stopwords:
                continue
            platform_tokens[platform].append(w_norm)
            global_counter[w_norm] += 1

    return platform_tokens, global_counter


def get_top_keywords(global_counter: Counter, top_n: int = 20) -> List[Tuple[str, int]]:
    """
    从全局词频中选出 Top N 关键词。
    """
    return global_counter.most_common(top_n)


def analyze_cross_platform_overlap(
    platform_tokens: Dict[str, List[str]]
) -> pd.DataFrame:
    """
    分析“跨平台重合度”，找出同时出现在 2 个或 3 个平台上的核心关键词。

    思路：
        1. 每个平台先把词列表转换为集合（去重），只关心“是否出现过”；
        2. 对所有平台进行遍历，统计每个词出现在哪些平台；
        3. 过滤掉只出现在 1 个平台的词；
        4. 返回一个 DataFrame，其中包含：
           - Keyword: 关键词
           - Platforms: 出现的平台（逗号分隔）
           - PlatformCount: 出现的平台数量（2 或 3）
    """
    # 1. 将每个平台的词列表去重成集合
    platform_word_sets: Dict[str, set] = {
        platform: set(words) for platform, words in platform_tokens.items()
    }

    # 2. 统计每个词在哪些平台出现
    word_platforms: Dict[str, List[str]] = defaultdict(list)
    for platform, word_set in platform_word_sets.items():
        for w in word_set:
            word_platforms[w].append(platform)

    rows = []
    for word, platforms in word_platforms.items():
        if len(platforms) <= 1:
            continue  # 只出现在一个平台，不算“跨平台共性话题”

        rows.append(
            {
                "Keyword": word,
                "Platforms": ",".join(sorted(platforms)),
                "PlatformCount": len(platforms),
            }
        )

    df = pd.DataFrame(rows, columns=["Keyword", "Platforms", "PlatformCount"])
    # 按照参与平台数量降序，再按关键词排序
    df = df.sort_values(by=["PlatformCount", "Keyword"], ascending=[False, True])
    return df


def build_word_stats_dataframe(
    global_counter: Counter,
    platform_tokens: Dict[str, List[str]],
    overlap_df: pd.DataFrame,
    top_n: int = 20,
) -> pd.DataFrame:
    """
    构造“词频与共现分析”用的 DataFrame（用于导出到 Excel 的第 5 个工作表）。

    字段示例：
        - Keyword
        - TotalFreq
        - WeiboFreq
        - DouyinFreq
        - BaiduFreq
        - Platforms
        - PlatformCount
    """
    # 先分别统计各个平台的词频
    platform_counters: Dict[str, Counter] = {
        platform: Counter(words) for platform, words in platform_tokens.items()
    }

    # 按全网词频取 Top N
    top_keywords = [word for word, _ in global_counter.most_common(top_n)]

    overlap_info = overlap_df.set_index("Keyword") if not overlap_df.empty else None

    rows = []
    for word in top_keywords:
        total_freq = global_counter[word]
        weibo_freq = platform_counters.get("微博", Counter()).get(word, 0)
        douyin_freq = platform_counters.get("抖音", Counter()).get(word, 0)
        baidu_freq = platform_counters.get("百度", Counter()).get(word, 0)
        bilibili_freq = platform_counters.get("哔哩哔哩", Counter()).get(word, 0)
        kr36_freq = platform_counters.get("36氪", Counter()).get(word, 0)

        if overlap_info is not None and word in overlap_info.index:
            platforms = overlap_info.loc[word, "Platforms"]
            platform_count = int(overlap_info.loc[word, "PlatformCount"])
        else:
            # 即便不跨平台，也记录下来（平台数=1，方便观察）
            platforms = ",".join(
                p for p, c in [
                    ("微博", weibo_freq),
                    ("抖音", douyin_freq),
                    ("百度", baidu_freq),
                ]
                if c > 0
            )
            platform_count = len(platforms.split(",")) if platforms else 0

        rows.append(
            {
                "Keyword": word,
                "TotalFreq": total_freq,
                "WeiboFreq": weibo_freq,
                "DouyinFreq": douyin_freq,
                "BaiduFreq": baidu_freq,
                "BilibiliFreq": bilibili_freq,
                "Kr36Freq": kr36_freq,
                "Platforms": platforms,
                "PlatformCount": platform_count,
            }
        )

    df = pd.DataFrame(
        rows,
        columns=[
            "Keyword",
            "TotalFreq",
            "WeiboFreq",
            "DouyinFreq",
            "BaiduFreq",
            "BilibiliFreq",
            "Kr36Freq",
            "Platforms",
            "PlatformCount",
        ],
    )
    return df


############################
# 模块四：多维图表可视化
############################

def plot_platform_top5_bar(all_df: pd.DataFrame, output_path: str) -> None:
    """
    图表 1：对比柱状图

    要求：
        - 展示微博 / 抖音 / 百度三大平台各自热度 Top 5 话题；
        - 以一张图的形式展示（这里采用 1 行 3 列子图）。
    """
    plt.figure(figsize=(18, 9))

    platforms = [p["name"] for p in PLATFORMS]

    # 2 行 3 列布局（适配最多 6 个平台；目前 5 个平台）
    for i, platform in enumerate(platforms, start=1):
        sub_df = all_df[all_df["Platform"] == platform].copy()
        if sub_df.empty:
            continue
        # 选取热度前 5
        sub_df = sub_df.sort_values(by="HeatScore", ascending=False).head(5)
        titles = sub_df["Title"].tolist()
        scores = sub_df["HeatScore"].tolist()

        ax = plt.subplot(2, 3, i)
        bars = ax.bar(range(len(titles)), scores)
        ax.set_title(f"{platform} 热度 Top 5")
        ax.set_xlabel("话题")
        ax.set_ylabel("热度指数")
        ax.set_xticks(range(len(titles)))

        # 使用換行來控制長度，保留完整詞語而不是只截取前幾個字
        import textwrap

        wrapped_labels = [
            "\n".join(textwrap.wrap(t, width=8)) if len(t) > 8 else t
            for t in titles
        ]
        ax.set_xticklabels(wrapped_labels, rotation=25, ha="right", fontsize=9)

        # 在柱子顶端标出数值
        for bar, score in zip(bars, scores):
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                height,
                f"{int(score):d}",
                ha="center",
                va="bottom",
                fontsize=8,
            )

    plt.suptitle("多平台 热度 Top 5 话题对比", fontsize=14)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])

    try:
        plt.savefig(output_path, dpi=200)
        logging.info("平台 Top5 对比柱状图已保存至: %s", output_path)
    except Exception as e:
        logging.error("保存平台 Top5 对比柱状图失败: %s", e)
    finally:
        plt.close()


def generate_wordcloud(global_counter: Counter, font_path: Optional[str], output_path: str) -> None:
    """
    图表 2：全网热词词云图。
    """
    if not global_counter:
        logging.warning("词频数据为空，无法生成词云。")
        return

    # WordCloud 需要传入 {词: 频次} 字典
    word_freq_dict = dict(global_counter)

    if not font_path:
        logging.warning("未找到可用的中文字体文件，词云中的中文可能会乱码。")

    wc = WordCloud(
        font_path=font_path or None,
        width=800,
        height=400,
        background_color="white",
        max_words=200,
    )

    try:
        wc.generate_from_frequencies(word_freq_dict)
        wc.to_file(output_path)
        logging.info("全网热词词云图已保存至: %s", output_path)
    except Exception as e:
        logging.error("生成或保存词云图失败: %s", e)


def plot_platform_top5_bar_base64(all_df: pd.DataFrame) -> Optional[str]:
    """
    生成平台 Top5 对比柱状图，返回 base64 編碼後的 PNG。
    用於在網頁中直接內嵌顯示。
    """
    if all_df.empty:
        return None

    fig = plt.figure(figsize=(18, 9))
    platforms = [p["name"] for p in PLATFORMS]

    for i, platform in enumerate(platforms, start=1):
        sub_df = all_df[all_df["Platform"] == platform].copy()
        if sub_df.empty:
            continue
        sub_df = sub_df.sort_values(by="HeatScore", ascending=False).head(5)
        titles = sub_df["Title"].tolist()
        scores = sub_df["HeatScore"].tolist()

        ax = fig.add_subplot(2, 3, i)
        bars = ax.bar(range(len(titles)), scores)
        ax.set_title(f"{platform} 热度 Top 5")
        ax.set_xlabel("话题")
        ax.set_ylabel("热度指数")
        ax.set_xticks(range(len(titles)))

        import textwrap

        wrapped_labels = [
            "\n".join(textwrap.wrap(t, width=8)) if len(t) > 8 else t
            for t in titles
        ]
        ax.set_xticklabels(wrapped_labels, rotation=25, ha="right", fontsize=9)

        for bar, score in zip(bars, scores):
            height = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                height,
                f"{int(score):d}",
                ha="center",
                va="bottom",
                fontsize=8,
            )

    fig.suptitle("多平台 热度 Top 5 话题对比", fontsize=14)
    fig.tight_layout(rect=[0, 0.03, 1, 0.95])

    buf = io.BytesIO()
    try:
        fig.savefig(buf, format="png", dpi=200)
        buf.seek(0)
        encoded = base64.b64encode(buf.read()).decode("ascii")
        return encoded
    except Exception as e:
        logging.error("生成平台 Top5 对比图的 base64 数据失败: %s", e)
        return None
    finally:
        plt.close(fig)
        buf.close()


def generate_wordcloud_base64(
    global_counter: Counter, font_path: Optional[str]
) -> Optional[str]:
    """
    生成全网热词词云圖，返回 base64 編碼後的 PNG。
    用於在網頁中直接內嵌顯示。
    """
    if not global_counter:
        return None

    word_freq_dict = dict(global_counter)

    wc = WordCloud(
        font_path=font_path or None,
        width=800,
        height=400,
        background_color="white",
        max_words=200,
    )

    try:
        wc.generate_from_frequencies(word_freq_dict)
        img = wc.to_image()
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        encoded = base64.b64encode(buf.read()).decode("ascii")
        return encoded
    except Exception as e:
        logging.error("生成词云图 base64 数据失败: %s", e)
        return None


############################
# 模块五：Excel 与简报导出
############################

def style_header_row(ws, fill: PatternFill) -> None:
    """
    为工作表的首行（表头）添加背景色。
    """
    for cell in ws[1]:
        cell.fill = fill


def add_heat_data_bar(ws, heat_col_name: str) -> None:
    """
    为指定列添加数据条条件格式（Data Bar）。

    参数:
        ws: openpyxl 工作表对象
        heat_col_name: 列名（如 "HeatScore" 或 "TotalFreq"）
    """
    # 在首行中找到该列的列号
    header_row = ws[1]
    target_col_idx = None
    for idx, cell in enumerate(header_row, start=1):
        if str(cell.value) == heat_col_name:
            target_col_idx = idx
            break

    if target_col_idx is None:
        return

    col_letter = get_column_letter(target_col_idx)
    # 从第 2 行开始到最后一行
    start_row = 2
    end_row = ws.max_row
    # 如果只有表头（没有数据行），则不添加条件格式
    if end_row < start_row:
        return
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

    rule = DataBarRule(
        start_type="min",
        end_type="max",
        color="63C384",  # 一种绿色
        showValue=True,
    )
    ws.conditional_formatting.add(cell_range, rule)


def export_to_excel(
    all_df: pd.DataFrame,
    weibo_df: pd.DataFrame,
    douyin_df: pd.DataFrame,
    baidu_df: pd.DataFrame,
    bilibili_df: pd.DataFrame,
    kr36_df: pd.DataFrame,
    word_stats_df: pd.DataFrame,
    output_path: str,
) -> None:
    """
    导出多工作表 Excel 分析报告。

    Sheet 1: 全网总榜
    Sheet 2: 微博热榜
    Sheet 3: 抖音热榜
    Sheet 4: 百度热榜
    Sheet 5: 哔哩哔哩热榜
    Sheet 6: 36氪热榜
    Sheet 7: 词频与共现分析
    """
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            all_df.to_excel(writer, sheet_name="全网总榜", index=False)
            weibo_df.to_excel(writer, sheet_name="微博热榜", index=False)
            douyin_df.to_excel(writer, sheet_name="抖音热榜", index=False)
            baidu_df.to_excel(writer, sheet_name="百度热榜", index=False)
            bilibili_df.to_excel(writer, sheet_name="哔哩哔哩热榜", index=False)
            kr36_df.to_excel(writer, sheet_name="36氪热榜", index=False)
            word_stats_df.to_excel(writer, sheet_name="词频与共现分析", index=False)

            # 获取 openpyxl 的 workbook 对象
            wb = writer.book

            # 不同工作表的表头填充颜色
            fill_total = PatternFill("solid", fgColor="CCCCCC")  # 灰色
            fill_weibo = PatternFill("solid", fgColor="FFEB9C")  # 浅黄色
            fill_douyin = PatternFill("solid", fgColor="C6EFCE")  # 浅绿色
            fill_baidu = PatternFill("solid", fgColor="BDD7EE")  # 浅蓝色
            fill_bili = PatternFill("solid", fgColor="D9E1F2")   # 淡紫蓝
            fill_kr36 = PatternFill("solid", fgColor="E2F0D9")   # 淡绿
            fill_word = PatternFill("solid", fgColor="F8CBAD")   # 浅橙色

            ws_total = wb["全网总榜"]
            ws_weibo = wb["微博热榜"]
            ws_douyin = wb["抖音热榜"]
            ws_baidu = wb["百度热榜"]
            ws_bili = wb["哔哩哔哩热榜"]
            ws_kr36 = wb["36氪热榜"]
            ws_word = wb["词频与共现分析"]

            # 冻结首行
            for ws in [ws_total, ws_weibo, ws_douyin, ws_baidu, ws_bili, ws_kr36, ws_word]:
                ws.freeze_panes = "A2"

            # 设置表头颜色
            style_header_row(ws_total, fill_total)
            style_header_row(ws_weibo, fill_weibo)
            style_header_row(ws_douyin, fill_douyin)
            style_header_row(ws_baidu, fill_baidu)
            style_header_row(ws_bili, fill_bili)
            style_header_row(ws_kr36, fill_kr36)
            style_header_row(ws_word, fill_word)

            # 为热度相关列添加数据条条件格式
            add_heat_data_bar(ws_total, "HeatScore")
            add_heat_data_bar(ws_weibo, "HeatScore")
            add_heat_data_bar(ws_douyin, "HeatScore")
            add_heat_data_bar(ws_baidu, "HeatScore")
            add_heat_data_bar(ws_bili, "HeatScore")
            add_heat_data_bar(ws_kr36, "HeatScore")
            add_heat_data_bar(ws_word, "TotalFreq")

        logging.info("Excel 报告已成功导出至: %s", output_path)
    except Exception as e:
        logging.error("导出 Excel 报告失败: %s", e)


def export_markdown_brief(
    overlap_df: pd.DataFrame,
    global_counter: Counter,
    output_path: str,
    top_k: int = 3,
) -> str:
    """
    自动生成“今日全网舆情简报.md”，总结全网最热的若干个跨平台共性话题。
    """
    today_str = datetime.now().strftime("%Y-%m-%d")

    if overlap_df.empty:
        content = (
            f"# 今日全网舆情简报（{today_str}）\n\n"
            "由于今天跨平台共性话题不足，未能计算出稳定的重合热点。\n"
            "建议检查数据源是否正常或适当放宽筛选条件。\n"
        )
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        logging.info("今日全网舆情简报已生成（但无跨平台共性话题）: %s", output_path)
        return content

    # 将 overlap 信息与全局词频结合，找出“最热的跨平台话题”
    tmp_df = overlap_df.copy()
    tmp_df["TotalFreq"] = tmp_df["Keyword"].apply(lambda w: global_counter.get(w, 0))
    tmp_df = tmp_df.sort_values(
        by=["PlatformCount", "TotalFreq"], ascending=[False, False]
    )
    top_df = tmp_df.head(top_k)

    lines = [
        f"# 今日全网舆情简报（{today_str}）",
        "",
        "以下为今日全网最受关注的跨平台共性话题（按覆盖平台数与热度综合排序）：",
        "",
    ]

    for idx, row in top_df.iterrows():
        keyword = row["Keyword"]
        platforms = row["Platforms"]
        platform_count = int(row["PlatformCount"])
        total_freq = int(row["TotalFreq"])

        lines.append(
            f"{len(lines) - 3}. **{keyword}**  "
            f"(出现平台：{platforms}，涉及平台数：{platform_count}，综合词频：{total_freq})"
        )

    lines.append("")
    lines.append("> 以上结果仅基于今日样本数据，实际舆情需结合更多上下文与时间维度进行研判。")

    content = "\n".join(lines)
    try:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        logging.info("今日全网舆情简报已生成: %s", output_path)
    except Exception as e:
        logging.error("生成 Markdown 简报失败: %s", e)
    return content


############################
# 主流程控制
############################

def build_history_record(
    all_df: pd.DataFrame,
    weibo_df: pd.DataFrame,
    douyin_df: pd.DataFrame,
    baidu_df: pd.DataFrame,
    bilibili_df: pd.DataFrame,
    kr36_df: pd.DataFrame,
    word_stats_df: pd.DataFrame,
    overlap_df: pd.DataFrame,
    top_keywords: List[Tuple[str, int]],
    request_status: Dict[str, Dict[str, Any]],
) -> Dict[str, Any]:
    """
    构建一次运行的历史记录（用于趋势分析）。
    只保存必要字段，避免文件过大。
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    record: Dict[str, Any] = {
        "generated_at": ts,
        "request_status": request_status,
        "counts": {
            "all": int(len(all_df)),
            "weibo": int(len(weibo_df)),
            "douyin": int(len(douyin_df)),
            "baidu": int(len(baidu_df)),
            "bilibili": int(len(bilibili_df)),
            "kr36": int(len(kr36_df)),
        },
        # 关键词：Top20
        "top_keywords": [{"keyword": k, "freq": int(v)} for k, v in top_keywords],
        # 共性话题：取前 50，便于趋势展示
        "overlap": overlap_df.head(50).to_dict(orient="records") if not overlap_df.empty else [],
        # 词频与共现分析：取前 50
        "word_stats": word_stats_df.head(50).to_dict(orient="records") if not word_stats_df.empty else [],
        # 各平台 Top1（按热度）
        "platform_top1": {
            "weibo": (
                weibo_df.sort_values("HeatScore", ascending=False).head(1).to_dict(orient="records")
                if not weibo_df.empty else []
            ),
            "douyin": (
                douyin_df.sort_values("HeatScore", ascending=False).head(1).to_dict(orient="records")
                if not douyin_df.empty else []
            ),
            "baidu": (
                baidu_df.sort_values("HeatScore", ascending=False).head(1).to_dict(orient="records")
                if not baidu_df.empty else []
            ),
            "bilibili": (
                bilibili_df.sort_values("HeatScore", ascending=False).head(1).to_dict(orient="records")
                if not bilibili_df.empty else []
            ),
            "kr36": (
                kr36_df.sort_values("HeatScore", ascending=False).head(1).to_dict(orient="records")
                if not kr36_df.empty else []
            ),
        },
    }
    return record


def save_history_record(record: Dict[str, Any]) -> Optional[str]:
    """
    将历史记录保存到 `history/` 目录，返回保存的文件路径。
    """
    try:
        out_dir = Path("history")
        out_dir.mkdir(parents=True, exist_ok=True)
        filename = datetime.now().strftime("history_%Y%m%d_%H%M%S.json")
        path = out_dir / filename
        with path.open("w", encoding="utf-8") as f:
            json.dump(record, f, ensure_ascii=False, indent=2)
        return str(path)
    except Exception as e:
        logging.error("保存运行历史失败: %s", e)
        return None


def load_recent_history(limit: int = 20) -> List[Dict[str, Any]]:
    """
    加载最近 N 次运行历史（按文件名时间倒序）。
    """
    out_dir = Path("history")
    if not out_dir.exists():
        return []
    files = sorted(out_dir.glob("history_*.json"), reverse=True)
    records: List[Dict[str, Any]] = []
    for fp in files[: max(1, limit)]:
        try:
            with fp.open("r", encoding="utf-8") as f:
                records.append(json.load(f))
        except Exception:
            continue
    return list(reversed(records))


def build_keyword_trend_series(
    history_records: List[Dict[str, Any]],
    keyword: str,
) -> Tuple[List[str], List[int]]:
    """
    从历史记录中提取某关键词的词频趋势。
    """
    xs: List[str] = []
    ys: List[int] = []
    kw = normalize_keyword(keyword)
    for r in history_records:
        ts = str(r.get("generated_at", ""))
        xs.append(ts)
        freq = 0
        try:
            for item in r.get("top_keywords", []):
                if normalize_keyword(item.get("keyword", "")) == kw:
                    freq = int(item.get("freq", 0))
                    break
        except Exception:
            freq = 0
        ys.append(freq)
    return xs, ys


def build_overlap_count_series(history_records: List[Dict[str, Any]]) -> Tuple[List[str], List[int]]:
    """
    跨平台共性话题数量趋势（每次运行 overlap 的条数）。
    """
    xs: List[str] = []
    ys: List[int] = []
    for r in history_records:
        xs.append(str(r.get("generated_at", "")))
        try:
            ys.append(int(len(r.get("overlap", []) or [])))
        except Exception:
            ys.append(0)
    return xs, ys


def build_platform_top1_heat_series(
    history_records: List[Dict[str, Any]],
    platform_key: str,
) -> Tuple[List[str], List[float]]:
    """
    平台 Top1 热度趋势（每次运行取该平台热榜 HeatScore 最大的那条）。

    platform_key: "weibo" / "douyin" / "baidu"
    """
    xs: List[str] = []
    ys: List[float] = []
    for r in history_records:
        xs.append(str(r.get("generated_at", "")))
        heat = 0.0
        try:
            top1_list = (((r.get("platform_top1") or {}).get(platform_key)) or [])
            if isinstance(top1_list, list) and top1_list:
                heat = float(top1_list[0].get("HeatScore", 0) or 0)
        except Exception:
            heat = 0.0
        ys.append(heat)
    return xs, ys


def plot_multi_line_trend_base64(
    xs: List[str],
    series: List[Tuple[str, List[float]]],
    title: str,
    y_label: str,
) -> Optional[str]:
    """
    绘制多条趋势折线图并返回 base64 PNG。
    """
    if not xs or not series:
        return None
    fig = plt.figure(figsize=(10, 3.8))
    ax = fig.add_subplot(1, 1, 1)

    x_idx = list(range(len(xs)))
    for name, ys in series:
        if len(ys) != len(xs):
            continue
        ax.plot(x_idx, ys, marker="o", label=name)

    ax.set_title(title)
    ax.set_xlabel("时间（最近运行）")
    ax.set_ylabel(y_label)
    labels = [x[5:16] if len(x) >= 16 else x for x in xs]
    ax.set_xticks(x_idx)
    ax.set_xticklabels(labels, rotation=25, ha="right", fontsize=8)
    ax.grid(True, linestyle="--", alpha=0.3)
    ax.legend(fontsize=9, loc="best")
    fig.tight_layout()

    buf = io.BytesIO()
    try:
        fig.savefig(buf, format="png", dpi=180)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("ascii")
    except Exception as e:
        logging.error("生成多折线趋势图失败: %s", e)
        return None
    finally:
        plt.close(fig)
        buf.close()


def plot_trend_line_base64(
    xs: List[str],
    ys: List[int],
    title: str,
) -> Optional[str]:
    """
    绘制趋势折线图并返回 base64 PNG。
    """
    if not xs or not ys or len(xs) != len(ys):
        return None
    fig = plt.figure(figsize=(10, 3.6))
    ax = fig.add_subplot(1, 1, 1)
    ax.plot(range(len(xs)), ys, marker="o")
    ax.set_title(title)
    ax.set_xlabel("时间（最近运行）")
    ax.set_ylabel("词频")
    # 只显示少量刻度，避免挤
    tick_idx = list(range(len(xs)))
    labels = [x[5:16] if len(x) >= 16 else x for x in xs]
    ax.set_xticks(tick_idx)
    ax.set_xticklabels(labels, rotation=25, ha="right", fontsize=8)
    ax.grid(True, linestyle="--", alpha=0.3)
    fig.tight_layout()

    buf = io.BytesIO()
    try:
        fig.savefig(buf, format="png", dpi=180)
        buf.seek(0)
        return base64.b64encode(buf.read()).decode("ascii")
    except Exception as e:
        logging.error("生成趋势图失败: %s", e)
        return None
    finally:
        plt.close(fig)
        buf.close()

def run_full_analysis() -> Optional[Dict[str, Any]]:
    """
    執行完整分析流程，並返回可供網頁界面使用的結果字典。

    返回的字典包含：
        - all_df / weibo_df / douyin_df / baidu_df
        - word_stats_df / overlap_df / top_keywords
        - global_counter
        - bar_chart_b64 / wordcloud_b64
        - brief_md  （今日全網輿情簡報內容）
        - font_path （實際使用的中文字體路徑）
    """
    setup_logging()

    # 配置字体
    font_path = get_available_font()
    setup_matplotlib(font_path)

    fetcher = HotRankFetcher(API_CONFIG)

    # 1. 获取数据
    raw_weibo = fetcher.fetch_weibo()
    raw_douyin = fetcher.fetch_douyin()
    raw_baidu = fetcher.fetch_baidu()
    raw_bilibili = fetcher.fetch_bilibili()
    raw_kr36 = fetcher.fetch_36kr()

    if not any([raw_weibo, raw_douyin, raw_baidu, raw_bilibili, raw_kr36]):
        logging.error("所有平台的数据全部获取失败，程序终止。")
        return None

    # 2. 数据清洗与统一
    weibo_df = normalize_records_from_json("微博", raw_weibo)
    douyin_df = normalize_records_from_json("抖音", raw_douyin)
    baidu_df = normalize_records_from_json("百度", raw_baidu)
    bilibili_df = normalize_records_from_json("哔哩哔哩", raw_bilibili)
    kr36_df = normalize_records_from_json("36氪", raw_kr36)

    if (
        weibo_df.empty
        and douyin_df.empty
        and baidu_df.empty
        and bilibili_df.empty
        and kr36_df.empty
    ):
        logging.error("所有平台经解析后的数据均为空，程序终止。")
        return None

    all_df = merge_all_platforms(weibo_df, douyin_df, baidu_df, bilibili_df, kr36_df)
    logging.info("多平台合并后的总数据量：%d 条", len(all_df))

    # 3. 文本分词与共现统计
    platform_tokens, global_counter = tokenize_titles_by_platform(all_df)
    top_keywords = get_top_keywords(global_counter, top_n=20)
    logging.info("全网 Top20 关键词：%s", top_keywords)

    overlap_df = analyze_cross_platform_overlap(platform_tokens)
    logging.info("跨平台共性话题数量：%d", len(overlap_df))

    word_stats_df = build_word_stats_dataframe(
        global_counter, platform_tokens, overlap_df, top_n=20
    )

    # 输出：保存一次“运行历史”，便于趋势分析
    history_record = build_history_record(
        all_df=all_df,
        weibo_df=weibo_df,
        douyin_df=douyin_df,
        baidu_df=baidu_df,
        bilibili_df=bilibili_df,
        kr36_df=kr36_df,
        word_stats_df=word_stats_df,
        overlap_df=overlap_df,
        top_keywords=top_keywords,
        request_status=fetcher.last_request_status,
    )
    history_path = save_history_record(history_record)

    # 4. 图表可视化（文件輸出）
    output_dir = os.path.abspath(".")
    bar_chart_path = os.path.join(output_dir, "Platform_Top5_Compare.png")
    wordcloud_path = os.path.join(output_dir, "Global_Hot_Words.png")

    plot_platform_top5_bar(all_df, bar_chart_path)
    generate_wordcloud(global_counter, font_path, wordcloud_path)

    # 4.1 圖像的 base64 版本（用於網頁顯示）
    bar_chart_b64 = plot_platform_top5_bar_base64(all_df)
    wordcloud_b64 = generate_wordcloud_base64(global_counter, font_path)

    # 5. 导出 Excel 报表与 Markdown 简报
    today_str = datetime.now().strftime("%Y%m%d")
    excel_path = os.path.join(
        output_dir, f"全网热搜综合分析报告_{today_str}.xlsx"
    )
    export_to_excel(
        all_df=all_df,
        weibo_df=weibo_df,
        douyin_df=douyin_df,
        baidu_df=baidu_df,
        bilibili_df=bilibili_df,
        kr36_df=kr36_df,
        word_stats_df=word_stats_df,
        output_path=excel_path,
    )

    md_path = os.path.join(output_dir, "今日全网舆情简报.md")
    brief_md = export_markdown_brief(
        overlap_df=overlap_df,
        global_counter=global_counter,
        output_path=md_path,
        top_k=3,
    )

    logging.info("程序执行完毕。")

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "font_path": font_path,
        "request_status": fetcher.last_request_status,
        "all_df": all_df,
        "weibo_df": weibo_df,
        "douyin_df": douyin_df,
        "baidu_df": baidu_df,
        "bilibili_df": bilibili_df,
        "kr36_df": kr36_df,
        "word_stats_df": word_stats_df,
        "overlap_df": overlap_df,
        "top_keywords": top_keywords,
        "global_counter": global_counter,
        "bar_chart_b64": bar_chart_b64,
        "wordcloud_b64": wordcloud_b64,
        "brief_md": brief_md,
        "output_dir": output_dir,
        "excel_filename": os.path.basename(excel_path),
        "bar_chart_filename": os.path.basename(bar_chart_path),
        "wordcloud_filename": os.path.basename(wordcloud_path),
        "md_filename": os.path.basename(md_path),
        "history_file": os.path.basename(history_path) if history_path else "",
        "synonym_count": int(len(SYNONYM_FLAT_MAP)),
    }


# ========== Web 可视化界面部分 ==========

app = Flask(__name__)

# 全局緩存一次分析結果，避免每次刷新都重新調用 API
ANALYSIS_RESULTS: Optional[Dict[str, Any]] = None
ANALYSIS_LAST_RUN_TS: Optional[float] = None
# 10 分鐘快取：老師檢查時刷新頁面不會重複打 API
CACHE_TTL_SECONDS = 600

HTML_TEMPLATE = """
<!doctype html>
<html lang="zh-CN">
<head>
    <meta charset="utf-8">
    <title>全网舆情与跨平台热点追踪分析系统</title>
    <style>
        body {
            font-family: "Microsoft YaHei", "SimHei", sans-serif;
            margin: 20px;
            background-color: #f9f9f9;
        }
        h1 { color: #333333; }
        h2 { color: #444444; margin-top: 30px; }
        h3 { color: #555555; margin-top: 20px; }
        .info {
            margin-bottom: 15px;
            color: #555;
            font-size: 14px;
        }
        img {
            max-width: 100%;
            height: auto;
            border: 1px solid #ddd;
            box-shadow: 0 2px 6px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }
        .table-container {
            overflow-x: auto;
            background-color: #ffffff;
            padding: 10px;
            border-radius: 4px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
            margin-bottom: 20px;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            font-size: 12px;
        }
        th, td {
            border: 1px solid #dddddd;
            padding: 4px 6px;
            text-align: left;
        }
        th {
            background-color: #f2f2f2;
        }
        .brief {
            white-space: pre-wrap;
            background-color: #ffffff;
            padding: 12px;
            border-radius: 4px;
            border: 1px solid #e0e0e0;
            font-size: 13px;
        }
        .footer {
            margin-top: 30px;
            font-size: 12px;
            color: #888888;
        }
        .controls {
            margin: 12px 0 18px 0;
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }
        .controls-group {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            align-items: center;
        }
        .btn {
            display: inline-block;
            padding: 8px 12px;
            border-radius: 4px;
            border: 1px solid #cccccc;
            background: #ffffff;
            color: #333333;
            text-decoration: none;
            font-size: 13px;
            cursor: pointer;
        }
        .btn.primary {
            border-color: #2f6fed;
            background: #2f6fed;
            color: white;
        }
        .btn.danger {
            border-color: #b42318;
            background: #b42318;
            color: white;
        }
        .badge {
            display: inline-block;
            padding: 2px 6px;
            border-radius: 10px;
            background: #eef2ff;
            color: #2f6fed;
            font-size: 12px;
            margin-left: 6px;
        }
        .status-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
            gap: 10px;
            margin: 12px 0 18px 0;
        }
        .status-card {
            background: #ffffff;
            border: 1px solid #e6e6e6;
            border-radius: 6px;
            padding: 10px 12px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
            font-size: 13px;
        }
        .ok { color: #1b7f3a; font-weight: 600; }
        .bad { color: #b42318; font-weight: 600; }

        .summary-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 10px;
            margin: 12px 0 18px 0;
        }
        .summary-card {
            background: #ffffff;
            border: 1px solid #e6e6e6;
            border-radius: 6px;
            padding: 10px 12px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
            font-size: 13px;
        }
        .summary-title { color: #666; font-size: 12px; }
        .summary-value { font-size: 18px; font-weight: 700; margin-top: 4px; }
        .search-panel {
            background: #ffffff;
            border: 1px solid #e6e6e6;
            border-radius: 6px;
            padding: 12px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
            margin: 12px 0 18px 0;
        }
        select, input[type="text"] {
            padding: 6px 8px;
            border-radius: 4px;
            border: 1px solid #cccccc;
            font-size: 13px;
        }
        .muted { color: #777; font-size: 12px; }
        mark {
            background: #fff2a8;
            padding: 0 2px;
            border-radius: 2px;
        }
    </style>
</head>
<body>
    <h1>全网舆情与跨平台热点追踪分析系统</h1>
    <div class="info">
        学生：高二丙 班　学号：16　姓名：冯俊帮<br>
        说明：本页面由 <b>Python 程序自动生成</b>，集中展示微博 / 抖音 / 百度三大平台的今日舆情分析结果。
    </div>

    <div class="controls">
        <div class="controls-group">
            <a class="btn primary" href="/refresh">重新抓取并分析（可能需 10~30 秒）</a>
            <a class="btn danger" href="/refresh?force=1">强制刷新（无视 10 分钟缓存）</a>
        </div>
        <div class="controls-group">
            <a class="btn" href="/download/{{ excel_filename }}">下载 Excel 报告</a>
            <a class="btn" href="/download/{{ bar_chart_filename }}">下载 Top5 对比图</a>
            <a class="btn" href="/download/{{ wordcloud_filename }}">下载 词云图</a>
            <a class="btn" href="/download/{{ md_filename }}">下载 Markdown 简报</a>
            <a class="btn" href="/export_html{{ current_query }}">导出静态 HTML 报告</a>
        </div>
    </div>

    <h2>零、运行状态与总览</h2>
    <div class="info">
        生成时间：<b>{{ generated_at }}</b>
        {% if keyword %}
            <span class="badge">已筛选关键词：{{ keyword }}</span>
        {% endif %}
        <span class="badge">同义词映射：{{ synonym_count }} 条</span>
    </div>

    <div class="summary-grid">
        <div class="summary-card">
            <div class="summary-title">全网总榜条目数</div>
            <div class="summary-value">{{ total_rows }}</div>
        </div>
        <div class="summary-card">
            <div class="summary-title">微博条目数</div>
            <div class="summary-value">{{ weibo_rows }}</div>
        </div>
        <div class="summary-card">
            <div class="summary-title">抖音条目数</div>
            <div class="summary-value">{{ douyin_rows }}</div>
        </div>
        <div class="summary-card">
            <div class="summary-title">百度条目数</div>
            <div class="summary-value">{{ baidu_rows }}</div>
        </div>
        <div class="summary-card">
            <div class="summary-title">哔哩哔哩条目数</div>
            <div class="summary-value">{{ bilibili_rows }}</div>
        </div>
        <div class="summary-card">
            <div class="summary-title">36氪条目数</div>
            <div class="summary-value">{{ kr36_rows }}</div>
        </div>
    </div>

    <div class="status-grid">
        {% for p in platforms %}
            {% set s = request_status.get(p.key) %}
            <div class="status-card">
                <div><b>{{ p.name }}</b></div>
                {% if s and s.ok %}
                    <div>状态：<span class="ok">成功</span>（{{ s.status_code }}）</div>
                    <div>耗时：{{ s.elapsed_ms }} ms　重试次数：{{ s.attempts }}</div>
                {% else %}
                    <div>状态：<span class="bad">失败</span></div>
                    <div>错误：{{ (s.error if s else "无状态信息") }}</div>
                    <div>耗时：{{ (s.elapsed_ms if s else "-") }} ms　尝试次数：{{ (s.attempts if s else "-") }}</div>
                {% endif %}
            </div>
        {% endfor %}
    </div>

    <div class="search-panel">
        <form method="get" action="/">
            <div class="controls-group">
                <label><b>关键词筛选（下拉可选）</b></label>
                <select name="keyword">
                    <option value="">（不筛选）</option>
                    {% for kw in keyword_options %}
                        <option value="{{ kw }}" {% if kw == keyword %}selected{% endif %}>{{ kw }}</option>
                    {% endfor %}
                </select>

                <label><b>匹配方式</b></label>
                <select name="match">
                    <option value="title" {% if match_mode == "title" %}selected{% endif %}>标题包含（直观）</option>
                    <option value="token" {% if match_mode == "token" %}selected{% endif %}>分词匹配（更严谨）</option>
                </select>

                <label><b>表格显示行数</b></label>
                <select name="rows">
                    <option value="10" {% if rows == 10 %}selected{% endif %}>10</option>
                    <option value="50" {% if rows == 50 %}selected{% endif %}>50</option>
                    <option value="all" {% if rows is none %}selected{% endif %}>全部</option>
                </select>

                <label><b>平台过滤</b></label>
                <select name="platform">
                    <option value="all" {% if platform_filter == "all" %}selected{% endif %}>全部</option>
                    <option value="weibo" {% if platform_filter == "weibo" %}selected{% endif %}>微博</option>
                    <option value="douyin" {% if platform_filter == "douyin" %}selected{% endif %}>抖音</option>
                    <option value="baidu" {% if platform_filter == "baidu" %}selected{% endif %}>百度</option>
                    <option value="bilibili" {% if platform_filter == "bilibili" %}selected{% endif %}>哔哩哔哩</option>
                    <option value="kr36" {% if platform_filter == "kr36" %}selected{% endif %}>36氪</option>
                </select>

                <label><b>排序</b></label>
                <select name="sort">
                    <option value="Rank" {% if sort_col == "Rank" %}selected{% endif %}>Rank</option>
                    <option value="HeatScore" {% if sort_col == "HeatScore" %}selected{% endif %}>HeatScore</option>
                </select>
                <select name="dir">
                    <option value="asc" {% if sort_dir == "asc" %}selected{% endif %}>升序</option>
                    <option value="desc" {% if sort_dir == "desc" %}selected{% endif %}>降序</option>
                </select>

                <label><b>TopN（按当前排序）</b></label>
                <select name="topn">
                    <option value="" {% if topn is none %}selected{% endif %}>不限制</option>
                    <option value="10" {% if topn == 10 %}selected{% endif %}>10</option>
                    <option value="20" {% if topn == 20 %}selected{% endif %}>20</option>
                    <option value="50" {% if topn == 50 %}selected{% endif %}>50</option>
                </select>

                <button class="btn primary" type="submit">应用</button>
                <a class="btn" href="/">清除筛选</a>
            </div>
            <div class="muted" style="margin-top:8px;">
                提示：选择关键词后，下方四个平台表格会自动筛选命中条目；词频表会置顶该关键词，并展示该关键词出现的平台。
            </div>
        </form>

        {% if keyword and keyword_platforms %}
            <div style="margin-top:10px;">
                <b>关键词出现平台：</b> {{ keyword_platforms | join(", ") }}
            </div>
        {% elif keyword %}
            <div style="margin-top:10px;">
                <b>关键词出现平台：</b> 未在本次样本标题中匹配到（可能是分词关键词，不一定直接出现在标题里）
            </div>
        {% endif %}
    </div>

    <h2>二、趋势追踪（最近 {{ history_count }} 次运行）</h2>
    <div class="muted">说明：每次运行都会自动保存到 `history/`。下方趋势图按你选择的关键词展示其在 Top20 里的词频变化。</div>
    {% if trend_b64 %}
        <img src="data:image/png;base64,{{ trend_b64 }}" alt="关键词趋势图">
    {% else %}
        <p class="muted">暂无趋势图（可能历史不足，或该关键词在历史 Top20 中未出现）。</p>
    {% endif %}

    <h3>1. 各平台 Top1 热度趋势</h3>
    {% if platform_top1_trend_b64 %}
        <img src="data:image/png;base64,{{ platform_top1_trend_b64 }}" alt="各平台 Top1 热度趋势图">
    {% else %}
        <p class="muted">暂无平台 Top1 热度趋势图（可能历史不足）。</p>
    {% endif %}

    <h3>2. 跨平台共性话题数量趋势</h3>
    {% if overlap_count_trend_b64 %}
        <img src="data:image/png;base64,{{ overlap_count_trend_b64 }}" alt="跨平台共性话题数量趋势图">
    {% else %}
        <p class="muted">暂无共性话题数量趋势图（可能历史不足）。</p>
    {% endif %}

    <h2>一、平台热度 Top 5 话题对比</h2>
    {% if bar_chart_b64 %}
        <img src="data:image/png;base64,{{ bar_chart_b64 }}" alt="平台 Top5 对比柱状图">
    {% else %}
        <p>暂无平台 Top5 图像数据。</p>
    {% endif %}

    <h2>二、全网热词词云</h2>
    {% if wordcloud_b64 %}
        <img src="data:image/png;base64,{{ wordcloud_b64 }}" alt="全网热词词云图">
    {% else %}
        <p>暂无词云数据。</p>
    {% endif %}

    <h2>三、全网与各平台热榜（节选）</h2>
    <h3>1. 全网总榜（{{ table_hint }}）</h3>
    <div class="muted">筛选后条目数：{{ all_rows_filtered }}</div>
    <div class="table-container">
        {{ all_df_html | safe }}
    </div>

    <h3>2. 微博热榜（{{ table_hint }}）</h3>
    <div class="muted">筛选后条目数：{{ weibo_rows_filtered }}</div>
    <div class="table-container">
        {{ weibo_df_html | safe }}
    </div>

    <h3>3. 抖音热榜（{{ table_hint }}）</h3>
    <div class="muted">筛选后条目数：{{ douyin_rows_filtered }}</div>
    <div class="table-container">
        {{ douyin_df_html | safe }}
    </div>

    <h3>4. 百度热榜（{{ table_hint }}）</h3>
    <div class="muted">筛选后条目数：{{ baidu_rows_filtered }}</div>
    <div class="table-container">
        {{ baidu_df_html | safe }}
    </div>

    <h3>5. 哔哩哔哩热榜（{{ table_hint }}）</h3>
    <div class="muted">筛选后条目数：{{ bilibili_rows_filtered }}</div>
    <div class="table-container">
        {{ bilibili_df_html | safe }}
    </div>

    <h3>6. 36氪热榜（{{ table_hint }}）</h3>
    <div class="muted">筛选后条目数：{{ kr36_rows_filtered }}</div>
    <div class="table-container">
        {{ kr36_df_html | safe }}
    </div>

    <h2>四、词频与跨平台共现分析（Top 20）</h2>
    <h3>1. 全网 Top 20 关键词词频</h3>
    <div class="table-container">
        {{ word_stats_html | safe }}
    </div>

    <h3>2. 跨平台共性话题（节选）</h3>
    <div class="table-container">
        {{ overlap_html | safe }}
    </div>

    <h3>3. 今日跨平台共性话题 Top 3（综合排序）</h3>
    <div class="table-container">
        {{ overlap_top3_html | safe }}
    </div>

    <h2>五、今日全网舆情简报</h2>
    <div class="brief">
{{ brief_md }}
    </div>

    <div class="footer">
        页面由 Python + Flask 自动生成，数据源：第三方公开热榜 API（微博 / 抖音 / 百度）。<br>
        如需重新分析，请在终端重新运行：<code>python main.py</code>。
    </div>
</body>
</html>
"""


@app.route("/")
def dashboard() -> str:
    """
    Web 主頁：顯示圖像、表格與簡報內容。
    """
    if ANALYSIS_RESULTS is None:
        return "<h3>分析结果尚未生成 (内存已重置)。</h3><p>请先回到 <a href='/'>首页前台展示</a>，找到 Live Demo 区域并点击【执行全网分析脚本】获取最新数据！</p>"

    r = ANALYSIS_RESULTS

    # 读取筛选参数：rows=10/50/all，keyword=下拉选择
    rows_raw = request.args.get("rows", "10")
    if rows_raw == "all":
        rows = None
    else:
        try:
            rows = int(rows_raw)
        except Exception:
            rows = 10

    keyword = (request.args.get("keyword") or "").strip()
    match_mode = (request.args.get("match") or "title").strip().lower()
    if match_mode not in {"title", "token"}:
        match_mode = "title"

    platform_filter = (request.args.get("platform") or "all").strip().lower()
    if platform_filter not in {"all", "weibo", "douyin", "baidu"}:
        platform_filter = "all"

    sort_col = (request.args.get("sort") or "Rank").strip()
    if sort_col not in {"Rank", "HeatScore"}:
        sort_col = "Rank"

    sort_dir = (request.args.get("dir") or "asc").strip().lower()
    if sort_dir not in {"asc", "desc"}:
        sort_dir = "asc"
    ascending = sort_dir == "asc"

    topn_raw = (request.args.get("topn") or "").strip()
    topn: Optional[int]
    if not topn_raw:
        topn = None
    else:
        try:
            topn = int(topn_raw)
            if topn <= 0:
                topn = None
        except Exception:
            topn = None

    def _limit_df(df: pd.DataFrame) -> pd.DataFrame:
        if rows is None:
            return df
        return df.head(max(1, rows))

    def _filter_by_keyword(df: pd.DataFrame) -> pd.DataFrame:
        if not keyword:
            return df
        if "Title" not in df.columns:
            return df
        title_series = df["Title"].astype(str)
        if match_mode == "title":
            # 在标题中做包含匹配（简单直观）
            mask = title_series.str.contains(keyword, na=False)
            return df[mask]

        # 分词匹配（更严谨）：对每条标题分词后，看是否包含该关键词
        stopwords = load_stopwords()
        kw_norm = normalize_keyword(keyword)
        def _hit(title: str) -> bool:
            try:
                words = jieba.lcut(title)
            except Exception:
                return False
            cleaned = []
            for w in words:
                w = str(w).strip()
                if not w or w in stopwords:
                    continue
                if w.isdigit():
                    continue
                if len(w) <= 1:
                    continue
                cleaned.append(normalize_keyword(w))
            return kw_norm in cleaned

        return df[title_series.apply(_hit)]

    def _highlight_keyword_in_title(df: pd.DataFrame) -> pd.DataFrame:
        """
        将 Title 中命中的 keyword 高亮（仅用于网页显示）。
        """
        if not keyword or "Title" not in df.columns:
            return df
        out = df.copy()
        try:
            out["Title"] = out["Title"].astype(str).apply(
                lambda t: t.replace(keyword, f"<mark>{keyword}</mark>")
            )
        except Exception:
            pass
        return out

    def _apply_sort_and_topn(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return df
        if sort_col in df.columns:
            try:
                df = df.sort_values(by=sort_col, ascending=ascending)
            except Exception:
                pass
        if topn is not None:
            df = df.head(topn)
        return df

    # 关键词可选项：Top20 + 跨平台共性词 + (可选) 全部词频前 200
    keyword_options: List[str] = []
    try:
        top20 = [w for w, _ in (r.get("top_keywords") or [])]
    except Exception:
        top20 = []
    overlap_words = []
    try:
        if isinstance(r.get("overlap_df"), pd.DataFrame) and not r["overlap_df"].empty:
            overlap_words = r["overlap_df"]["Keyword"].astype(str).head(200).tolist()
    except Exception:
        overlap_words = []

    # 从全局词频中补充一些（避免选项太少）
    global_words = []
    try:
        gc = r.get("global_counter")
        if gc:
            global_words = [w for w, _ in gc.most_common(200)]
    except Exception:
        global_words = []

    seen = set()
    for w in (top20 + overlap_words + global_words):
        w = str(w).strip()
        if not w or w in seen:
            continue
        seen.add(w)
        keyword_options.append(w)

    # 计算关键词出现平台（基于标题包含 或 分词匹配）
    keyword_platforms: List[str] = []
    if keyword:
        for p in PLATFORMS:
            platform_name = p["name"]
            df = r.get(f"{p['key']}_df")
            if not isinstance(df, pd.DataFrame):
                continue
            try:
                if df.empty:
                    continue
                if match_mode == "title":
                    if df["Title"].astype(str).str.contains(keyword, na=False).any():
                        keyword_platforms.append(platform_name)
                else:
                    stopwords = load_stopwords()
                    kw_norm = normalize_keyword(keyword)
                    def _hit_any(t: str) -> bool:
                        try:
                            words = jieba.lcut(t)
                        except Exception:
                            return False
                        cleaned = []
                        for w in words:
                            w = str(w).strip()
                            if not w or w in stopwords:
                                continue
                            if w.isdigit():
                                continue
                            if len(w) <= 1:
                                continue
                            cleaned.append(normalize_keyword(w))
                        return kw_norm in cleaned
                    if df["Title"].astype(str).apply(_hit_any).any():
                        keyword_platforms.append(platform_name)
            except Exception:
                pass

    # 先对各平台按“排序 + TopN”处理
    all_base = _apply_sort_and_topn(r["all_df"])
    weibo_base = _apply_sort_and_topn(r["weibo_df"])
    douyin_base = _apply_sort_and_topn(r["douyin_df"])
    baidu_base = _apply_sort_and_topn(r["baidu_df"])

    # 平台过滤：影响“全网总榜”显示（其余平台表仍独立展示，便于对照）
    if platform_filter == "weibo":
        all_base = all_base[all_base["Platform"] == "微博"]
    elif platform_filter == "douyin":
        all_base = all_base[all_base["Platform"] == "抖音"]
    elif platform_filter == "baidu":
        all_base = all_base[all_base["Platform"] == "百度"]

    # 表格数据：按关键词过滤、再按行数截取
    all_df_filtered = _filter_by_keyword(all_base)
    weibo_df_filtered = _filter_by_keyword(weibo_base)
    douyin_df_filtered = _filter_by_keyword(douyin_base)
    baidu_df_filtered = _filter_by_keyword(baidu_base)
    bilibili_df_filtered = _filter_by_keyword(_apply_sort_and_topn(r.get("bilibili_df", pd.DataFrame())))
    kr36_df_filtered = _filter_by_keyword(_apply_sort_and_topn(r.get("kr36_df", pd.DataFrame())))

    all_df_view = _limit_df(_highlight_keyword_in_title(all_df_filtered))
    weibo_df_view = _limit_df(_highlight_keyword_in_title(weibo_df_filtered))
    douyin_df_view = _limit_df(_highlight_keyword_in_title(douyin_df_filtered))
    baidu_df_view = _limit_df(_highlight_keyword_in_title(baidu_df_filtered))
    bilibili_df_view = _limit_df(_highlight_keyword_in_title(bilibili_df_filtered))
    kr36_df_view = _limit_df(_highlight_keyword_in_title(kr36_df_filtered))

    all_df_html = all_df_view.to_html(
        classes="dataframe", index=False, border=0
        , escape=False
    )
    weibo_df_html = weibo_df_view.to_html(
        classes="dataframe", index=False, border=0
        , escape=False
    )
    douyin_df_html = douyin_df_view.to_html(
        classes="dataframe", index=False, border=0
        , escape=False
    )
    baidu_df_html = baidu_df_view.to_html(
        classes="dataframe", index=False, border=0
        , escape=False
    )
    bilibili_df_html = bilibili_df_view.to_html(
        classes="dataframe", index=False, border=0
        , escape=False
    )
    kr36_df_html = kr36_df_view.to_html(
        classes="dataframe", index=False, border=0
        , escape=False
    )

    # 词频表：若选了关键词，将该行置顶（更像“搜索结果”）
    word_stats_df: pd.DataFrame = r["word_stats_df"].copy()
    if keyword and "Keyword" in word_stats_df.columns:
        try:
            hit = word_stats_df[word_stats_df["Keyword"].astype(str) == keyword]
            rest = word_stats_df[word_stats_df["Keyword"].astype(str) != keyword]
            word_stats_df = pd.concat([hit, rest], ignore_index=True)
        except Exception:
            pass

    word_stats_html = word_stats_df.head(20).to_html(
        classes="dataframe", index=False, border=0
    )
    overlap_html = (
        r["overlap_df"]
        .head(20)
        .to_html(classes="dataframe", index=False, border=0)
    )

    # Top3 共性话题：按平台覆盖数 & 全局词频综合排序
    overlap_top3_html = "<p>暂无跨平台共性话题</p>"
    try:
        odf: pd.DataFrame = r["overlap_df"].copy()
        if not odf.empty:
            gc = r.get("global_counter") or Counter()
            odf["TotalFreq"] = odf["Keyword"].apply(lambda w: int(gc.get(w, 0)))
            top3 = odf.sort_values(by=["PlatformCount", "TotalFreq"], ascending=[False, False]).head(3)
            overlap_top3_html = top3.to_html(classes="dataframe", index=False, border=0)
    except Exception:
        pass

    table_hint = f"前 {rows} 行" if rows is not None else "全部"

    # 趋势图：最近 N 次历史中，该关键词在 Top20 的频次变化
    history_records = load_recent_history(limit=20)
    history_count = len(history_records)
    trend_b64 = None
    if keyword and history_records:
        xs, ys = build_keyword_trend_series(history_records, keyword)
        trend_b64 = plot_trend_line_base64(xs, ys, title=f"关键词趋势：{keyword}")

    # 趋势图：各平台 Top1 热度趋势（多折线）
    platform_top1_trend_b64 = None
    overlap_count_trend_b64 = None
    if history_records:
        xs_h = [str(r0.get("generated_at", "")) for r0 in history_records]
        # 平台 Top1 热度多折线（动态适配平台数量）
        series = []
        xs_ok = True
        for p in PLATFORMS:
            xs_p, ys_p = build_platform_top1_heat_series(history_records, p["key"])
            if xs_p != xs_h:
                xs_ok = False
                break
            series.append((p["name"], ys_p))
        if xs_ok and series:
            platform_top1_trend_b64 = plot_multi_line_trend_base64(
                xs_h,
                series=series,
                title="各平台 Top1 热度趋势",
                y_label="HeatScore",
            )

        # 共性话题数量趋势
        xs_o, ys_o = build_overlap_count_series(history_records)
        if xs_o == xs_h:
            overlap_count_trend_b64 = plot_trend_line_base64(
                xs_o, ys_o, title="跨平台共性话题数量趋势"
            )

    # 维持当前 query，用于“导出静态 HTML”
    current_query = ""
    try:
        if request.query_string:
            current_query = "?" + request.query_string.decode("utf-8", errors="ignore")
    except Exception:
        current_query = ""

    return render_template_string(
        HTML_TEMPLATE,
        bar_chart_b64=r.get("bar_chart_b64"),
        wordcloud_b64=r.get("wordcloud_b64"),
        all_df_html=all_df_html,
        weibo_df_html=weibo_df_html,
        douyin_df_html=douyin_df_html,
        baidu_df_html=baidu_df_html,
        bilibili_df_html=bilibili_df_html,
        kr36_df_html=kr36_df_html,
        word_stats_html=word_stats_html,
        overlap_html=overlap_html,
        overlap_top3_html=overlap_top3_html,
        brief_md=r.get("brief_md", ""),
        generated_at=r.get("generated_at", ""),
        request_status=r.get("request_status", {}),
        excel_filename=r.get("excel_filename", ""),
        bar_chart_filename=r.get("bar_chart_filename", ""),
        wordcloud_filename=r.get("wordcloud_filename", ""),
        md_filename=r.get("md_filename", ""),
        history_file=r.get("history_file", ""),
        synonym_count=r.get("synonym_count", 0),
        rows=rows,
        keyword=keyword,
        match_mode=match_mode,
        keyword_options=keyword_options,
        keyword_platforms=keyword_platforms,
        table_hint=table_hint,
        total_rows=int(len(r["all_df"])),
        weibo_rows=int(len(r["weibo_df"])),
        douyin_rows=int(len(r["douyin_df"])),
        baidu_rows=int(len(r["baidu_df"])),
        bilibili_rows=int(len(r.get("bilibili_df", pd.DataFrame()))),
        kr36_rows=int(len(r.get("kr36_df", pd.DataFrame()))),
        all_rows_filtered=int(len(all_df_filtered)),
        weibo_rows_filtered=int(len(weibo_df_filtered)),
        douyin_rows_filtered=int(len(douyin_df_filtered)),
        baidu_rows_filtered=int(len(baidu_df_filtered)),
        bilibili_rows_filtered=int(len(bilibili_df_filtered)),
        kr36_rows_filtered=int(len(kr36_df_filtered)),
        platforms=PLATFORMS,
        platform_filter=platform_filter,
        sort_col=sort_col,
        sort_dir=sort_dir,
        topn=topn,
        trend_b64=trend_b64,
        history_count=history_count,
        current_query=current_query,
        platform_top1_trend_b64=platform_top1_trend_b64,
        overlap_count_trend_b64=overlap_count_trend_b64,
    )


@app.route("/refresh")
def refresh() -> str:
    """
    重新抓取與分析（帶快取 TTL，避免過度頻繁調用 API）。
    """
    global ANALYSIS_RESULTS, ANALYSIS_LAST_RUN_TS

    force = request.args.get("force", "0") == "1"
    now = time.time()
    if (
        (not force)
        and ANALYSIS_LAST_RUN_TS is not None
        and (now - ANALYSIS_LAST_RUN_TS) < CACHE_TTL_SECONDS
        and ANALYSIS_RESULTS is not None
    ):
        # 快取仍有效，直接返回首頁
        return dashboard()

    results = run_full_analysis()
    if not results:
        return "刷新失败：分析过程出错，请查看终端日志。"
    ANALYSIS_RESULTS = results
    ANALYSIS_LAST_RUN_TS = now
    return dashboard()


@app.route("/download/<path:filename>")
def download_file(filename: str):
    """
    下载导出文件（Excel / PNG / Markdown）。
    只允许下载 output_dir 下的文件，避免目录穿越风险。
    """
    if ANALYSIS_RESULTS is None:
        abort(404)

    output_dir = ANALYSIS_RESULTS.get("output_dir")
    if not output_dir or not os.path.isdir(output_dir):
        abort(404)

    # 只允许下载我们生成的那几个文件名
    allowed = {
        ANALYSIS_RESULTS.get("excel_filename"),
        ANALYSIS_RESULTS.get("bar_chart_filename"),
        ANALYSIS_RESULTS.get("wordcloud_filename"),
        ANALYSIS_RESULTS.get("md_filename"),
    }
    if filename not in allowed:
        abort(404)

    return send_from_directory(output_dir, filename, as_attachment=True)


@app.route("/export_html")
def export_html():
    """
    导出当前页面为静态 HTML 文件（便于老师离线打开查看）。
    会保留当前 query（筛选/排序/TopN 等）。
    """
    if ANALYSIS_RESULTS is None:
        abort(404)

    # 复用 dashboard 的渲染结果（包含当前 query 参数）
    html = dashboard()
    if not isinstance(html, str):
        html = str(html)

    output_dir = ANALYSIS_RESULTS.get("output_dir") or os.path.abspath(".")
    try:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        filename = datetime.now().strftime("report_%Y%m%d_%H%M%S.html")
        path = Path(output_dir) / filename
        with path.open("w", encoding="utf-8") as f:
            f.write(html)
    except Exception:
        abort(500)

    return send_from_directory(output_dir, filename, as_attachment=True)


def main() -> None:
    """
    主入口函數：
        1. 執行完整數據分析流程；
        2. 啟動本地 Flask 服務並自動打開瀏覽器顯示分析界面。
    """
    global ANALYSIS_RESULTS, ANALYSIS_LAST_RUN_TS

    results = run_full_analysis()
    if not results:
        logging.error("分析過程失敗，未啟動可視化界面。")
        return

    ANALYSIS_RESULTS = results
    ANALYSIS_LAST_RUN_TS = time.time()

    # 自動在預設瀏覽器中打開頁面
    def open_browser() -> None:
        try:
            webbrowser.open("http://127.0.0.1:5000", new=1)
        except Exception as e:
            logging.error("自動打開瀏覽器失敗: %s", e)

    threading.Timer(1.5, open_browser).start()
    logging.info("即將啟動本地可視化界面：http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=False)


if __name__ == "__main__":
    main()

